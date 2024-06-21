import time

import cv2

from ultralytics import YOLO

import prediction
from prediction import get_predict
"""
import dataset_collector
import yolov8
from prediction import get_predict
from roboflow import Roboflow
import threading
import time
from ultralytics import YOLO
import openpyxl
import app
from openpyxl.styles import Alignment
from copy import copy
import csv
"""
"""
#Capture data. input = capture_image_number. None = 200
dataset_collector.capture_img(15)

#Upload data. upload all images in collected_data folder
dataset_collector.upload_img()

#Download data. input = version. None = 3
yolov8.download_datasets(3)

#Train model. input = epochs. None = 50
yolov8.train_model(100)         

"""
#model = YOLO('runs/detect/latest_train_led/weights/best.pt')
#get_predict(model,"Screenshot 2024-05-09 163927.png")

# zoomed_image = cv2.resize(img, None, fx=2, fy=2, interpolation=cv2.INTER_LINEAR)
# start_row = int((zoomed_image.shape[0] - 480) / 2)
# start_col = int((zoomed_image.shape[1] - 640) / 2)
# cropped_image = zoomed_image[start_row:start_row + 480, start_col:start_col + 640]

#Create new file
"""
wb = openpyxl.Workbook()
sheet = wb.active
title = ['TIME', 'MODE', 'SPEED', 'VALUE']
for i, value in enumerate(title, start=1):
    sheet.cell(row=1, column=i, value=value)
wb.save("./run_history/1.xlsx")
"""
"""
#Write next line
wb = openpyxl.load_workbook('./run_history/1.xlsx')

# Chọn sheet bạn muốn ghi vào
sheet = wb.active  # Bạn cũng có thể chọn sheet bằng cách sử dụng wb['tên_sheet']

# Dữ liệu bạn muốn ghi
new_data = ['Dữ liệu cột 1', 'Dữ liệu cột 2', 'Dữ liệu cột 3']

# Xác định hàng tiếp theo trống
next_row = sheet.max_row + 1

# Ghi dữ liệu vào hàng tiếp theo
for i, value in enumerate(new_data, start=1):
    sheet.cell(row=next_row, column=i, value=value)
for column_cells in sheet.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    sheet.column_dimensions[column_cells[0].column_letter].width = length + 2
# Lưu lại tệp
wb.save('./run_history/1.xlsx')
"""
"""
result_dir = "run_history"
filename = result_dir + "/" + app.get_current_date() + ".csv"
header = ['TIME', 'MODE', 'SPEED', 'VALUE']
with open(filename, 'w', newline='') as file:
    writer = csv.writer(file)
    writer.writerow(header)
print(f"Tạo tệp '{filename}' và ghi dòng tiêu đề thành công!")
"""

"""
def test_camera():
    # Mở camera
    cap = cv2.VideoCapture(1)

    # Kiểm tra xem camera có được mở không
    if not cap.isOpened():
        print("Không thể mở camera.")
        return

    # Vòng lặp để hiển thị hình ảnh từ camera
    while True:
        # Đọc frame từ camera
        ret, frame = cap.read()

        # Kiểm tra xem frame có được đọc thành công không
        if not ret:
            print("Không thể đọc frame từ camera.")
            break

        # Hiển thị frame
        cv2.imshow('Camera', frame)

        # Đợi 1 miligiây, nhấn phím 'q' để thoát
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    # Giải phóng camera và đóng cửa sổ
    cap.release()
    cv2.destroyAllWindows()


# Gọi hàm để thực hiện kiểm tra camera
test_camera()
"""
model_led = YOLO('runs/detect/latest_train_led/weights/best.pt')
model_led.to('cuda')
while True:
    img = cv2.imread("Picture1.png")
    print(time.time())
    value = prediction.get_predict(model_led, img, False)
    print(time.time())
    print("==================================")
